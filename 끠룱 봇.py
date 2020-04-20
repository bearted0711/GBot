import discord
import asyncio
import time
import random
import datetime
import os
import logging
import urllib
import openpyxl
#import pynacl
#import youtube_dl
from discord.ext import commands
from discord.utils import get
#from bs4 import BeautifulSoup


app = commands.Bot(command_prefix='끠룱')


access_token = os.environ["BOT_TOKEN"]

calcResult = 0

@app.event
async def say_after(delay, what):
    await asyncio.sleep(delay)
    print(what)

async def main():
    print(f"started at {time.strftime('%X')}")
    await say_after(1, 'hello')
    await say_after(2, 'world')

    print(f"finished at {time.strftime('%X')}")

asyncio.run(main())
@app.event
async def on_ready():
    print("다음으로 로그인합니다 : ")
    print(app.user.name)
    print(app.user.id)
    print("봇 초대 : https://discordapp.com/oauth2/authorize?client_id=696388012438716427&scope=bot")
    print("정상 가동 완료")
    print(f"봇 가동 : {time.strftime('%X')}")
    print("====================================================================")


    await app.change_presence(activity=discord.Game(name="테드 말 안듣고 하고싶은대로 사려고",type=1))

@app.command(pass_context=True)
async def randomNum(ctx, num1, num2):
    picked = random.randint(int(num1), int(num2))
    await ctx.send('끠룱이가 뽑은 숫자는 : '+str(picked))

#@commands.has_permissions(administrator=True)
@app.command(name="추방", pass_context=True)
async def _kick(ctx, *, user_name: discord.Member, reason=None):
    await user_name.kick(reason=reason)
    await ctx.send(str(user_name)+"잘가라!")

#@commands.has_permissions(administrator=True)
@app.command(name="밴", pass_context=True)
async def _ban(ctx, *, user_name: discord.Member):
    await user_name.ban()
    await ctx.send(str(user_name)+"영원히 죽어라!")

#@commands.has_permissions(administrator=True)
@app.command(name="언밴", pass_context=True)
async def _unban(ctx, *, user_name):
    banned_users = await ctx.guild.bans()
    member_name, member_discriminator = user_name.split('#')
    for ban_entry in banned_users:
        user = ban_entry.user
        if (user.name, user.discriminator) == (member_name, member_discriminator):
            await ctx.guild.unban(user)
            await ctx.send(f"{user.mention}다시 돌아와")
            return
    
@app.command(name="새벽내놔", pass_context=True)
async def _HumanRole(ctx, member: discord.Member=None):
    member = member or ctx.message.author
    await member.add_roles(get(ctx.guild.roles, name="새벽러"))
    await ctx.channel.send(str(member)+"끠룱")

@app.command(name="청소", pass_context=True)
async def _clear(ctx, *, amount=5):
    await ctx.channel.purge(limit=amount)

@app.command(name="들어와", pass_context=True)
async def _join(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
        await channel.connect()
        await ctx.send("끠룱!")
    else:
        await ctx.send("너 거기 안들어가있잖아!")
@app.command(name="나가")
async def _leave(ctx):
    await app.voice_clients[0].disconnect()
    await ctx.send("끠룱!")

@app.command(name="블로그검색")
async def _search_blog(ctx, *, search_query):
    temp = 0
    url_base = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query="
    url = url_base + urllib.parse.quote(search_query)
    title = ["", "", ""] # 더 많은 검색 : 빈칸("")을 늘리셔야 합니다.
    link = ["", "", ""] # 더 많은 검색 : 빈칸("")을 늘리셔야 합니다.
    soup = BeautifulSoup(urllib.request.urlopen(url).read(), 'html.parser')
    result = soup.find_all('a', "sh_blog_title _sp_each_url _sp_each_title")
    embed = discord.Embed(title="검색 결과", description=" ", color=0x00ff56)
    for n in result:
        if temp == 3: # 더 많은 검색 : 숫자(3)를 늘리셔야 합니다.
            break
        title[temp] = n.get("title")
        link[temp] = n.get("href")
        embed.add_field(name=title[temp], value=link[temp], inline=False)
        temp+=1
    embed.set_footer(text="검색 완료!")
    await ctx.send(embed=embed)



@app.event
async def on_message(message):
    await app.process_commands(message)
    if message.author.bot:
        return None
    if message.content == "끠":
        await message.channel.send("룱")
    if message.content == "나":
        await message.channel.send("비보벳따우")
    if message.content == "보":
        await message.channel.send("보벳띠")
    if message.content == "종":
        await message.channel.send("로로 갈까요")
    if message.content == "명":
        await message.channel.send("동으로 갈까요")
    if message.content == "던질까":
        await message.channel.send("말까")
    if message.content == "던":
        await message.channel.send("던")
        await message.channel.send("던")
        await message.channel.send("던")
        await message.channel.send("던져!")
        await message.channel.send("던져!")
    
    if message.content.startswith("끠룱1부터10"):
        for x in range(10):
            await message.channel.send(x+1)

    if message.content == "끠룱":
        await message.channel.send (random.choice(["끠?","룱?","끠룱?","끠","룱","끠룱","끠룱!","끠!","룱!"]))
        print("끠룱")

    if message.content.startswith("끠룱계산"):
        print("끠룱계산")
        global calcResult
        param = message.content.split()
        try:
            if param[1].startswith("더하기"):
                calcResult = int(param[2])+int(param[3])
                await message.channel.send("끠룱 : "+str(calcResult))
            if param[1].startswith("빼기"):
                calcResult = int(param[2])-int(param[3])
                await message.channel.send("끠룱 : "+str(calcResult))
            if param[1].startswith("곱하기"):
                calcResult = int(param[2])*int(param[3])
                await message.channel.send("끠룱 : "+str(calcResult))
            if param[1].startswith("나누기"):
                calcResult = int(param[2])/int(param[3])
                await message.channel.send("끠룱 : "+str(calcResult))
        except IndexError:
            await message.channel.send("숫자를 안줬잖아!")
        except ValueError:
            await message.channel.send("숫자로 줘!")
        except ZeroDivisionError:
            await message.channel.send("You can't divide with 0.")

    if message.content.startswith("!끠룱기억") and not message.content.startswith("!끠룱지워"):
        file = openpyxl.load_workbook("끠룱.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("끠룱!")
                file.save("끠룱.xlsx")
                break

    if message.content.startswith("!끠룱알려"):
        file = openpyxl.load_workbook("끠룱.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break

    if message.content.startswith("!끠룱지워"):
        file = openpyxl.load_workbook("끠룱.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = "-"
                await message.channel.send("끠룱!")
                file.save("끠룱.xlsx")
                break

    if message.content.startswith("끠룱Embed실행"):
        embed=discord.Embed(title="Example Embed", description="이것은 Embed입니다.", color=0x00ff56)
        embed.set_author(name="저자의 이름", url="https://blog.naver.com/naver.com", icon_url="https://cdn.discordapp.com/attachments/642926226704564239/696387598087749722/Screenshot_20200406-005533_Gallery.jpg")
        embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/642926226704564239/696387598087749722/Screenshot_20200406-005533_Gallery.jpg")
        embed.add_field(name="이것은 필드입니다.", value="필드의 값입니다.", inline=True)
        embed.add_field(name="이것은 필드 2입니다.", value="필드의 값입니다.", inline=True)
        embed.add_field(name="이것은 필드 3입니다.", value="필드의 값입니다.", inline=True)
        embed.add_field(name="이것은 필드 4입니다.", value="필드의 값입니다.", inline=True)
        embed.set_footer(text="이것은 푸터입니다.")
        await message.channel.send(embed=embed)


app.run(access_token)
